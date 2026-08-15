from html import escape


def _read_csv_rows(src_path):
    import csv

    with open(src_path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def _rows_to_txt(rows):
    lines = []
    for row in rows:
        vals = ["" if v is None else str(v) for v in row]
        lines.append("\t".join(vals).rstrip("\t"))
    return "\n".join(lines) + "\n"


def _rows_to_html(rows):
    body = "".join(
        "<tr>" + "".join(f"<td>{escape('' if v is None else str(v))}</td>" for v in row) + "</tr>"
        for row in rows
    )
    return (
        "<!DOCTYPE html>\n<html><head><meta charset='utf-8'><title>Sheet</title>"
        "<style>table{border-collapse:collapse}td,th{border:1px solid #999;padding:4px 8px}</style>"
        f"</head><body><table>{body}</table></body></html>\n"
    )


def xlsx_to_csv(src_path, out_path):
    import csv

    from openpyxl import load_workbook

    wb = load_workbook(src_path, read_only=True, data_only=True)
    ws = wb.active
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
    wb.close()


def xlsx_to_txt(src_path, out_path):
    from openpyxl import load_workbook

    wb = load_workbook(src_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(_rows_to_txt(rows))


def xlsx_to_html(src_path, out_path):
    from openpyxl import load_workbook

    wb = load_workbook(src_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(_rows_to_html(rows))


def csv_to_xlsx(src_path, out_path):
    from openpyxl import Workbook

    rows = _read_csv_rows(src_path)
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(out_path)


def csv_to_txt(src_path, out_path):
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(_rows_to_txt(_read_csv_rows(src_path)))


def csv_to_html(src_path, out_path):
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(_rows_to_html(_read_csv_rows(src_path)))


HANDLERS = {
    ("XLSX", "CSV"): xlsx_to_csv,
    ("XLSX", "TXT"): xlsx_to_txt,
    ("XLSX", "HTML"): xlsx_to_html,
    ("CSV", "XLSX"): csv_to_xlsx,
    ("CSV", "TXT"): csv_to_txt,
    ("CSV", "HTML"): csv_to_html,
}
